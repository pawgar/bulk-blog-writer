"""Parsowanie content planu z pliku XLSX."""

import openpyxl


def read_xlsx_headers(filepath: str) -> list[str]:
    """Czyta nagłówki (pierwszy wiersz) z pliku XLSX. Zwraca listę stringów."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    wb.close()
    return headers


def auto_detect_columns(headers: list[str]) -> dict[str, int | None]:
    """Próbuje automatycznie zmapować kolumny. Zwraca dict z indeksami lub None."""
    headers_lower = [h.lower() for h in headers]
    col_map: dict[str, int | None] = {
        "title": None,
        "main_kw": None,
        "secondary_kw": None,
        "notes": None,
        "domain": None,
        "lang": None,
    }

    for i, h in enumerate(headers_lower):
        # Tytuł
        if col_map["title"] is None and any(
            k in h
            for k in [
                "tytuł", "tytul", "title", "temat", "nagłówek", "naglowek",
                "heading", "topic",
            ]
        ):
            col_map["title"] = i
        # Główne słowo kluczowe
        elif col_map["main_kw"] is None and (
            (any(k in h for k in ["główne", "glowne", "main", "primary"])
             and any(k in h for k in ["słowo", "slowo", "keyword", "kw", "fraza"]))
            or h in ["główne kw", "glowne kw", "main kw", "main keyword",
                      "keyword", "słowo kluczowe", "slowo kluczowe", "fraza kluczowa"]
        ):
            col_map["main_kw"] = i
        # Słowa kluczowe poboczne
        elif col_map["secondary_kw"] is None and any(
            k in h
            for k in [
                "poboczne", "secondary", "supporting", "dodatkowe kw",
                "dodatkowe słowa", "dodatkowe slowa", "related", "long tail",
                "long-tail", "frazy poboczne", "słowa poboczne", "slowa poboczne",
            ]
        ):
            col_map["secondary_kw"] = i
        # Dodatkowe informacje
        elif col_map["notes"] is None and any(
            k in h
            for k in [
                "dodatkowe informacje", "dodatkowe info", "additional",
                "uwagi", "notes", "notatki", "komentarz", "opis",
                "wskazówki", "wskazowki", "instructions", "brief",
            ]
        ):
            col_map["notes"] = i
        # Domena
        elif col_map["domain"] is None and any(
            k in h for k in ["domena", "domain", "strona", "website", "site", "url"]
        ):
            col_map["domain"] = i
        # Język
        elif col_map["lang"] is None and any(
            k in h for k in ["język", "jezyk", "lang", "language"]
        ):
            col_map["lang"] = i

    return col_map


def parse_content_plan(filepath: str, col_map: dict[str, int | None]) -> list[dict]:
    """Parsuje XLSX z podanym mapowaniem kolumn. Zwraca listę dict z artykułami."""
    if col_map.get("title") is None:
        raise ValueError("Nie wskazano kolumny z tytułem wpisu.")

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    articles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title_idx = col_map["title"]
        if title_idx >= len(row):
            continue
        title = str(row[title_idx] or "").strip()
        if not title:
            continue

        def _get(key: str) -> str:
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx] or "").strip()

        articles.append(
            {
                "title": title,
                "main_kw": _get("main_kw"),
                "secondary_kw": _get("secondary_kw"),
                "notes": _get("notes"),
                "domain": _get("domain"),
                "lang": _get("lang"),
            }
        )

    wb.close()
    return articles
